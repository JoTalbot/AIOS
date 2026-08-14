
def _send_trading_tg_alert(message: str) -> bool:
    """Отправка фоновых уведомлений Quant Radar отключена по запросу владельца."""
    return False

def _disabled_send_trading_tg_alert(message: str) -> bool:
    """Отправка уведомления о сделках квант-трейдинга в Telegram."""
    import urllib.request
    import json
    import os
    from pathlib import Path
    
    from tg_bot.credentials import read_systemd_credential, secret_from_env_or_credential

    token = secret_from_env_or_credential(
        "TELEGRAM_BOT_TOKEN", "AIOS_TELEGRAM_TOKEN", credential="telegram_token"
    )
    chat_id = os.environ.get("TELEGRAM_CHAT_ID") or read_systemd_credential(
        "telegram_owner_chat_id"
    )
    
    if not token or not chat_id:
        env_file = Path("/root/AIOS/.env")
        if env_file.exists():
            for line in env_file.read_text(encoding="utf-8").splitlines():
                if line.startswith("TELEGRAM_BOT_TOKEN=") and not token:
                    token = line.split("=", 1)[1].strip().strip('"').strip("'")
                elif line.startswith("TELEGRAM_CHAT_ID=") and not chat_id:
                    chat_id = line.split("=", 1)[1].strip().strip('"').strip("'")
                    
    if not token or not chat_id:
        return False
        
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": int(chat_id),
        "text": message,
        "parse_mode": "HTML"
    }
    try:
        req = urllib.request.Request(
            url,
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"}
        )
        with urllib.request.urlopen(req, timeout=10):
            return True
    except Exception:
        return False

"""
AIOS Quantitative Trading Engine & Signal Radar
Модуль количественного трейдинга, анализа сигналов и бумажной торговли (Paper Trading) AIOS.

ФУНКЦИОНАЛ:
1. Запрос живых котировок BTC/USDT, ETH/USDT, SOL/USDT, MATIC/USDT с ведущих бирж.
2. Расчет количественных индикаторов (SMA 5/20 Crossover, RSI 14, Bollinger Bands).
3. Генерация торговых сигналов (BUY_LONG, SELL_SHORT, HOLD) с уровнем уверенности (Confidence Score).
4. Безопасная симуляция бумажной торговли (Paper Trading) с фиксированным начальным балансом $1,000.00.
5. Интегрированный симулятор торгов на бирже Kraken с виртуальным балансом $100.00.
6. Встроенный Kill-Switch при просадке > 5.0%.
"""

import os
import json
import time
import math
import logging
import urllib.request
from typing import Dict, Any, List, Optional
from pathlib import Path

from aios_core.crypto_wallet import AIOSWalletManager, PUBLIC_RPC_NODES
from aios_core.quant_directional_policy import DirectionalV2Config
from aios_core.quant_report_formatters import (
    format_backtest_report,
    format_kraken_demo_report,
    format_portfolio_advice_report,
    format_positions_only_report,
    format_single_asset_analysis,
    format_unified_crypto_earnings_report,
)

# --- Консультирующие ML/RL-сигналы (read-only, без автоторговли) ---
def _load_ai_signals() -> dict:
    """Собирает консультирующие сигналы ML (catboost) и RL (PPO)."""
    out = {"ml": {}, "rl": {}, "available": False}
    try:
        from aios_core.quant.ml_signal_bridge import MLSignalBridge
        out["ml"] = MLSignalBridge().summary()
    except Exception as e:
        out["ml"] = {"error": str(e)[:100]}
    try:
        from aios_core.quant.rl_signal_bridge import RLSignalBridge
        out["rl"] = RLSignalBridge().summary()
    except Exception as e:
        out["rl"] = {"error": str(e)[:100]}
    out["available"] = bool(out["ml"].get("available")) or bool(out["rl"].get("available"))
    return out

def get_ai_signal_summary() -> dict:
    """Объединённый консультативный отчёт ML+RL."""
    return _load_ai_signals()


logger = logging.getLogger("AIOS.QuantTrading")

_ADVISORY_CACHE = {"loaded_at": 0.0, "ml": {}, "rl": {}}

def _advisory_signal_maps(ttl_seconds: int = 300):
    """Load ML/RL signal maps once per process/TTL, not once per symbol."""
    now = time.time()
    if now - float(_ADVISORY_CACHE["loaded_at"]) < ttl_seconds:
        return _ADVISORY_CACHE["ml"], _ADVISORY_CACHE["rl"]
    ml_map, rl_map = {}, {}
    try:
        from aios_core.quant.ml_signal_bridge import MLSignalBridge
        ml_map = {item["symbol"]: item for item in MLSignalBridge().all_signals()}
    except Exception:
        pass
    try:
        from aios_core.quant.rl_signal_bridge import RLSignalBridge
        rl_map = {item.get("asset"): item for item in RLSignalBridge().run_all().get("signals", []) if item.get("ok")}
    except Exception:
        pass
    _ADVISORY_CACHE.update({"loaded_at": now, "ml": ml_map, "rl": rl_map})
    return ml_map, rl_map


class MarketDataFeed:
    """Модуль забора живых рыночных котировок криптовалют."""

    @staticmethod
    def fetch_live_price(symbol: str = "BTCUSDT") -> Dict[str, Any]:
        """Запрашивает живую цену пары с публичных API Binance/Bitstamp/CoinGecko."""
        url = f"https://api.binance.com/api/v3/ticker/24hr?symbol={symbol.upper()}"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "AIOS-Quant-Engine/1.0"})
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                return {
                    "symbol": symbol.upper(),
                    "price": float(data.get("lastPrice", 0.0)),
                    "high_24h": float(data.get("highPrice", 0.0)),
                    "low_24h": float(data.get("lowPrice", 0.0)),
                    "volume_24h": float(data.get("volume", 0.0)),
                    "price_change_pct": float(data.get("priceChangePercent", 0.0)),
                    "source": "binance_live",
                    "timestamp": time.time()
                }
        except Exception as e:
            logger.warning(f"⚠️ Binance API недоступен, fallback запрос: {e}")

        # Fallback simulated live prices
        fallback_prices = {
            "BTCUSDT": 65420.0,
            "ETHUSDT": 3450.0,
            "SOLUSDT": 155.0,
            "MATICUSDT": 0.52
        }
        p = fallback_prices.get(symbol.upper(), 100.0)
        return {
            "symbol": symbol.upper(),
            "price": p,
            "high_24h": p * 1.02,
            "low_24h": p * 0.98,
            "volume_24h": 15000.0,
            "price_change_pct": 1.5,
            "source": "quant_fallback",
            "timestamp": time.time()
        }


class QuantSignalEngine:
    """Двигатель количественного анализа и индикаторов (SMA, RSI)."""

    def __init__(self, data_dir: str = "/root/AIOS/data"):
        # Умное разрешение путей (Docker/Host)
        is_docker = os.path.exists('/.dockerenv') or (os.path.exists('/proc/self/cgroup') and 'docker' in open('/proc/self/cgroup').read())
        if is_docker and os.path.exists("/app/data"):
            data_dir = "/app/data"
            
        self.data_dir = Path(data_dir)
        self.history_file = self.data_dir / "price_history_quant.json"
        self._ensure_file()

    def _ensure_file(self):
        self.data_dir.mkdir(parents=True, exist_ok=True)
        if not self.history_file.exists():
            with open(self.history_file, "w", encoding="utf-8") as f:
                json.dump({}, f)

    def load_history(self) -> Dict[str, List[float]]:
        try:
            with open(self.history_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    def save_history(self, history: Dict[str, List[float]]):
        with open(self.history_file, "w", encoding="utf-8") as f:
            json.dump(history, f, indent=2)

    _FUNDING_CACHE = {"timestamp": 0.0, "rates": {}}

    @classmethod
    def fetch_funding_rate(cls, asset_symbol: str) -> float:
        """Запрашивает все ставки фандинга (Funding Rate) одним пакетом за 0.2с с кэшированием."""
        now = time.time()
        if now - cls._FUNDING_CACHE["timestamp"] > 120 or not cls._FUNDING_CACHE["rates"]:
            try:
                url = "https://fapi.binance.com/fapi/v1/premiumIndex"
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=4) as resp:
                    data = json.loads(resp.read().decode())
                    rates_map = {}
                    for item in data:
                        sym = item.get("symbol", "").replace("USDT", "")
                        rates_map[sym] = float(item.get("lastFundingRate", 0.0)) * 100.0
                    cls._FUNDING_CACHE = {"timestamp": now, "rates": rates_map}
            except Exception:
                pass

        clean_sym = asset_symbol.upper()
        for p in ["KRAKEN_", "BINANCE_", "BYBIT_", "OKX_", "UNISWAP_V3_"]:
            clean_sym = clean_sym.replace(p, "")
        for s in ["USDT", "USDC", "USD"]:
            if clean_sym.endswith(s) and len(clean_sym) > len(s):
                clean_sym = clean_sym[:-len(s)]

        return cls._FUNDING_CACHE["rates"].get(clean_sym, 0.005)

    def record_and_analyze(self, symbol: str, current_price: float) -> Dict[str, Any]:
        """Записывает котировку и рассчитывает 360-градусные индикаторы: SMA, RSI, Bollinger, MACD, Funding, Orderbook Depth, News Sentiment."""
        history = self.load_history()
        prices = history.get(symbol, [])
        prices.append(current_price)

        # Храним последние 50 свечей
        prices = prices[-50:]
        history[symbol] = prices
        self.save_history(history)

        # 1. Расчет скользящих средних SMA (Fast 3, Slow 10)
        fast_period = min(3, len(prices))
        slow_period = min(10, len(prices))
        sma_fast = sum(prices[-fast_period:]) / fast_period
        sma_slow = sum(prices[-slow_period:]) / slow_period

        # 2. Расчет RSI (14)
        rsi = 50.0
        if len(prices) >= 5:
            gains = [max(prices[i] - prices[i-1], 0) for i in range(1, len(prices))]
            losses = [max(prices[i-1] - prices[i], 0) for i in range(1, len(prices))]
            avg_gain = sum(gains[-14:]) / min(14, len(gains))
            avg_loss = sum(losses[-14:]) / min(14, len(losses))

            if avg_loss > 0:
                rs = avg_gain / avg_loss
                rsi = 100.0 - (100.0 / (1.0 + rs))
            else:
                rsi = 100.0

        # 3. Расчет Bollinger Bands (20)
        period_bb = min(20, len(prices))
        sma_bb = sum(prices[-period_bb:]) / period_bb
        variance = sum((p - sma_bb) ** 2 for p in prices[-period_bb:]) / period_bb
        std_dev = math.sqrt(variance)
        upper_bb = sma_bb + (2.0 * std_dev)
        lower_bb = sma_bb - (2.0 * std_dev)

        # 4. Расчет MACD (Fast 12, Slow 26)
        p12 = prices[-min(12, len(prices)):]
        p26 = prices[-min(26, len(prices)):]
        ema12 = sum(p12) / len(p12)
        ema26 = sum(p26) / len(p26)
        macd_line = ema12 - ema26

        # 5. Ставка фандинга деривативов (Funding Rate)
        funding_rate = self.fetch_funding_rate(symbol)

        # 6. Глубина стакана ордеров (Orderbook Imbalance)
        orderbook_status = "BALANCED"
        try:
            from aios_core.orderbook_analyzer import AIOSOrderbookAnalyzer
            ob_info = AIOSOrderbookAnalyzer.analyze_orderbook(symbol)
            orderbook_status = ob_info.get("status", "BALANCED")
        except Exception:
            pass

        # 7. Комплексный скоринг сигналов
        bullish_score = 0
        bearish_score = 0
        reasons = []

        if current_price <= lower_bb:
            bullish_score += 2
            reasons.append(f"Касание нижней Боллинджера (${lower_bb:.2f})")
        elif current_price >= upper_bb:
            bearish_score += 2
            reasons.append(f"Пробой верхней Боллинджера (${upper_bb:.2f})")

        if rsi < 35.0:
            bullish_score += 2
            reasons.append(f"Перепроданность RSI ({rsi:.1f})")
        elif rsi > 65.0:
            bearish_score += 2
            reasons.append(f"Перекупленность RSI ({rsi:.1f})")

        if sma_fast > sma_slow:
            bullish_score += 1
            reasons.append("Бычье SMA (3 > 10)")
        elif sma_fast < sma_slow:
            bearish_score += 1
            reasons.append("Медвежье SMA (3 < 10)")

        if macd_line > 0:
            bullish_score += 1
        elif macd_line < 0:
            bearish_score += 1

        if orderbook_status == "BUY_WALL_SUPPORT":
            bullish_score += 1
            reasons.append("Поддержка стены покупателей в стакане")
        elif orderbook_status == "SELL_WALL_RESISTANCE":
            bearish_score += 1
            reasons.append("Сопротивление стены продавцов в стакане")

        if funding_rate < 0:
            bullish_score += 1
            reasons.append(f"Short Squeeze риск (Funding {funding_rate:.4f}%)")
        elif funding_rate > 0.03:
            bearish_score += 1
            reasons.append(f"Long Squeeze риск (Funding {funding_rate:.4f}%)")

        # ML/RL консультирующий фактор (cached per process/TTL)
        ml_prob_up = None
        rl_position = None
        _sym = symbol.replace("KRAKEN_", "").split("/")[0].upper()
        _mls, _rls = _advisory_signal_maps()
        _ml_sig = _mls.get(_sym)
        if _ml_sig:
            ml_prob_up = float(_ml_sig.get("prob_up", 0.5))
            if ml_prob_up >= 0.65:
                bullish_score += 1
                reasons.append(f"ML бычий (prob_up {ml_prob_up:.2f})")
            elif ml_prob_up <= 0.35:
                bearish_score += 1
                reasons.append(f"ML медвежий (prob_up {ml_prob_up:.2f})")
        _rl_sig = _rls.get(_sym)
        if _rl_sig:
            rl_position = float(_rl_sig.get("position", 0.5))
            if rl_position > 0.7:
                bullish_score += 1
                reasons.append(f"RL LONG (pos {rl_position})")
            elif rl_position < 0.3:
                bearish_score += 1
                reasons.append(f"RL FLAT/шорт (pos {rl_position})")

        # Принятие решения
        signal = "HOLD"
        confidence = 0.50
        reason_text = "Индикаторы в нейтральной зоне"

        if bullish_score >= 3 and bullish_score > bearish_score:
            signal = "BUY_LONG"
            confidence = min(0.99, 0.70 + (bullish_score * 0.06))
            reason_text = " + ".join(reasons)
        elif bearish_score >= 3 and bearish_score > bullish_score:
            signal = "SELL_SHORT"
            confidence = min(0.99, 0.70 + (bearish_score * 0.06))
            reason_text = " + ".join(reasons)

        return {
            "symbol": symbol,
            "current_price": current_price,
            "sma_fast": round(sma_fast, 2),
            "sma_slow": round(sma_slow, 2),
            "rsi": round(rsi, 1),
            "macd": round(macd_line, 4),
            "funding_rate": round(funding_rate, 4),
            "orderbook_status": orderbook_status,
            "signal": signal,
            "confidence": round(confidence, 2),
            "ml_prob_up": ml_prob_up,
            "rl_position": rl_position,
            "reason": reason_text,
        }


class PaperTradingSimulator:
    """Симулятор бумажной торговли с учетом комиссий и рисков."""

    # Консервативная оценка taker fee на каждой стороне сделки.
    FEE_RATE = 0.0015
    MAX_OPEN_POSITIONS = 5
    MIN_CASH_RESERVE_PCT = 0.30

    def __init__(self, data_dir: str = "/root/AIOS/data", portfolio_filename: str = "paper_portfolio.json", initial_balance: float = 1000.0):
        # Умное разрешение путей (Docker/Host)
        is_docker = os.path.exists('/.dockerenv') or (os.path.exists('/proc/self/cgroup') and 'docker' in open('/proc/self/cgroup').read())
        if is_docker and os.path.exists("/app/data"):
            data_dir = "/app/data"
            
        self.data_dir = Path(data_dir)
        self.portfolio_file = self.data_dir / portfolio_filename
        self.wallet = AIOSWalletManager(data_dir)
        self.initial_balance = initial_balance
        self._ensure_file()

    def _ensure_file(self):
        self.data_dir.mkdir(parents=True, exist_ok=True)
        if not self.portfolio_file.exists():
            default_port = {
                "initial_balance_usd": self.initial_balance,
                "cash_usd": self.initial_balance,
                "realized_pnl_usd": 0.0,
                "total_trades": 0,
                "winning_trades": 0,
                "positions": {}
            }
            with open(self.portfolio_file, "w", encoding="utf-8") as f:
                json.dump(default_port, f, indent=2)

    def load_portfolio(self) -> Dict[str, Any]:
        try:
            with open(self.portfolio_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    def save_portfolio(self, port: Dict[str, Any]):
        with open(self.portfolio_file, "w", encoding="utf-8") as f:
            json.dump(port, f, indent=2, ensure_ascii=False)

    def execute_paper_signal(self, signal_info: Dict[str, Any], is_kraken: bool = False) -> Dict[str, Any]:
        """Исполняет бумажную сделку на основе полученного количественного сигнала."""
        symbol = signal_info["symbol"]
        price = signal_info["current_price"]
        signal = signal_info["signal"]

        port = self.load_portfolio()
        positions = port.get("positions", {})
        pos = positions.get(symbol)

        trade_res = {"executed": False, "details": "Нет условий для сделки"}

        # Лимитируем объем одной сделки (20% от кэша)
        max_trade_usd = 200.0 if not is_kraken else 20.0 # Для Кракен-лимита в $100 сделка равна $20

        # 1. Покупка (BUY_LONG)
        if signal == "BUY_LONG" and not pos:
            max_positions = self.MAX_OPEN_POSITIONS if is_kraken else 5
            cash_reserve = (
                port.get("initial_balance_usd", self.initial_balance) * self.MIN_CASH_RESERVE_PCT
            )
            if len(positions) >= max_positions or port.get("cash_usd", 0.0) <= cash_reserve:
                return {
                    "trade": {"executed": False, "details": "Риск-фильтр: лимит позиций или резерв cash"},
                    "portfolio_summary": {
                        "cash_usd": round(port.get("cash_usd", 0.0), 2),
                        "open_positions": len(positions),
                    },
                }
            buy_amount_usd = min(port["cash_usd"] * 0.20, max_trade_usd)
            if buy_amount_usd >= 2.0:
                fee_usd = buy_amount_usd * self.FEE_RATE
                asset_qty = (buy_amount_usd - fee_usd) / price
                port["cash_usd"] -= buy_amount_usd
                positions[symbol] = {
                    "side": "LONG",
                    "entry_price": price,
                    "qty": asset_qty,
                    "invested_usd": buy_amount_usd,
                    "entry_fee_usd": round(fee_usd, 8),
                    "opened_at": time.time()
                }
                port["positions"] = positions
                port["total_trades"] += 1
                self.save_portfolio(port)

                trade_res = {
                    "executed": True,
                    "action": "OPEN_LONG",
                    "symbol": symbol,
                    "price": price,
                    "qty": round(asset_qty, 6),
                    "invested_usd": round(buy_amount_usd, 2)
                }
                logger.info(f"📈 [Paper Trading {'Kraken' if is_kraken else 'Binance'}] Открыта позиция LONG {symbol}: {asset_qty:.6f} по ${price:.2f}")
                _send_trading_tg_alert(f"📈 <b>[Quant Radar] Открыта позиция LONG</b>\n• Биржа: <b>{'Kraken' if is_kraken else 'Binance'}</b>\n• Пара: <code>{symbol}</code>\n• Цена входа: <b>${price:.2f}</b>\n• Объем: {asset_qty:.6f} (${buy_amount_usd:.2f})")

        # 2. Закрытие позиции, Take-Profit (+2%) и Trailing Stop-Loss (-1%)
        elif pos:
            entry_price = pos["entry_price"]
            qty = pos["qty"]
            invested = pos["invested_usd"]
            current_value = qty * price
            exit_fee_usd = current_value * self.FEE_RATE
            net_current_value = current_value - exit_fee_usd
            pnl_usd = net_current_value - invested
            pnl_pct = (pnl_usd / invested) * 100.0
            
            # Обновляем максимальную зафиксированную цену для трейлинга
            max_seen = pos.get("max_price_seen", entry_price)
            if price > max_seen:
                pos["max_price_seen"] = price
                max_seen = price

            # Условия закрытия:
            # a) Автоматический Take-Profit: +2.0% прибыли
            # b) Защитный Trailing Stop-Loss: -1.0% от входа или -1.2% от локального пика
            # c) Медвежий сигнал: SELL_SHORT
            should_close = False
            close_reason = ""
            if pnl_pct >= 2.0:
                should_close = True
                close_reason = f"🎯 TAKE-PROFIT (+{pnl_pct:.2f}%)"
            elif pnl_pct <= -1.0:
                should_close = True
                close_reason = f"🛑 STOP-LOSS ({pnl_pct:.2f}%)"
            elif max_seen > entry_price * 1.01 and price <= max_seen * 0.988:
                should_close = True
                close_reason = f"📉 TRAILING STOP (откат от пика ${max_seen:.2f})"
            elif signal == "SELL_SHORT":
                should_close = True
                close_reason = "🔴 Сигнал SELL_SHORT (медвежье пересечение SMA/RSI)"

            if should_close:
                port["cash_usd"] += net_current_value
                port["realized_pnl_usd"] += pnl_usd
                del positions[symbol]
                port["positions"] = positions

                if pnl_usd > 0:
                    port["winning_trades"] += 1
                    logger.info(f"🏆 [Paper Trading {'Kraken' if is_kraken else 'Binance'}] {close_reason}: +${pnl_usd:.2f} USD")
                else:
                    logger.info(f"🛡 [Paper Trading {'Kraken' if is_kraken else 'Binance'}] {close_reason}: ${pnl_usd:.2f} USD")

                self.save_portfolio(port)

                trade_res = {
                    "executed": True,
                    "action": "CLOSE_LONG",
                    "reason": close_reason,
                    "symbol": symbol,
                    "entry_price": entry_price,
                    "exit_price": price,
                    "pnl_usd": round(pnl_usd, 2),
                    "pnl_pct": round(pnl_pct, 2),
                    "fees_usd": round(pos.get("entry_fee_usd", 0.0) + exit_fee_usd, 8)
                }
                logger.info(f"📉 [Paper Trading {'Kraken' if is_kraken else 'Binance'}] Закрыта позиция {symbol}: PnL = ${pnl_usd:.2f} ({pnl_pct:.2f}%) | {close_reason}")
                icon = "🎯" if pnl_usd > 0 else "🛑"
                _send_trading_tg_alert(f"{icon} <b>[Quant Radar] Закрыта позиция {symbol}</b>\n• Биржа: <b>{'Kraken' if is_kraken else 'Binance'}</b>\n• Причина: <b>{close_reason}</b>\n• Вход: ${entry_price:.2f} ➔ Выход: ${price:.2f}\n• Результат (PnL): <b>{'+' if pnl_usd>0 else ''}${pnl_usd:.2f} USD ({pnl_pct:.2f}%)</b>")

        win_rate = (port["winning_trades"] / port["total_trades"] * 100) if port["total_trades"] > 0 else 0.0

        return {
            "trade": trade_res,
            "portfolio_summary": {
                "cash_usd": round(port["cash_usd"], 2),
                "realized_pnl_usd": round(port["realized_pnl_usd"], 2),
                "total_trades": port["total_trades"],
                "win_rate_pct": round(win_rate, 1),
                "open_positions": len(positions)
            }
        }


class QuantMasterOrchestrator:
    """Главный координатор количественного анализа и трейдинга AIOS."""

    def __init__(self, data_dir: str = "/root/AIOS/data"):
        # Умное разрешение путей (Docker/Host)
        is_docker = os.path.exists('/.dockerenv') or (os.path.exists('/proc/self/cgroup') and 'docker' in open('/proc/self/cgroup').read())
        if is_docker and os.path.exists("/app/data"):
            data_dir = "/app/data"
            
        self.signal_engine = QuantSignalEngine(data_dir)
        
        # Симулятор #1: Стандартный бумажный трейдинг Binance (Баланс $1,000)
        self.binance_simulator = PaperTradingSimulator(data_dir, "paper_portfolio.json", initial_balance=1000.0)
        
        # Симулятор #2: Интегрированный бумажный трейдинг Kraken (Баланс $100)
        self.kraken_simulator = PaperTradingSimulator(data_dir, "kraken_paper_portfolio.json", initial_balance=100.0)

    def run_quant_cycle(self) -> Dict[str, Any]:
        """Запуск цикла: Запрос котировок -> Анализ индикаторов -> Симулирование сделок."""
        logger.info("📈 [QuantEngine] Запуск цикла количественного анализа котировок...")

        # 1. Цикл 1: Трейдинг-симулятор Binance
        symbols = ["BTCUSDT", "ETHUSDT", "SOLUSDT"]
        binance_signals = []
        binance_trades = []

        for sym in symbols:
            market_data = MarketDataFeed.fetch_live_price(sym)
            analysis = self.signal_engine.record_and_analyze(sym, market_data["price"])
            trade_exec = self.binance_simulator.execute_paper_signal(analysis, is_kraken=False)

            binance_signals.append(analysis)
            binance_trades.append(trade_exec)

        # 2. Цикл 2: Трейдинг-симулятор KRAKEN (Виртуальный баланс $100)
        # Мы запрашиваем ЖИВЫЕ котировки непосредственно с API Кракен!
        from aios_core.kraken_client import AIOSKrakenClient
        kraken_client = AIOSKrakenClient()
        
        kraken_pairs_map = {
            "BTCUSD": "XXBTZUSD",
            "ETHUSD": "XETHZUSD",
            "SOLUSD": "SOLUSD",
            "XRPUSD": "XXRPZUSD",
            "ADAUSD": "ADAUSD",
            "DOTUSD": "DOTUSD",
            "LINKUSD": "LINKUSD",
            "AVAXUSD": "AVAXUSD",
            "LTCUSD": "XLTCZUSD",
            "NEARUSD": "NEARUSD",
            "UNIUSD": "UNIUSD",
            "SHIBUSD": "SHIBUSD",
            "DOGEUSD": "XDGUSD",
            "POLUSD": "POLUSD",
            "ATOMUSD": "ATOMUSD",
            "XLMUSD": "XXLMZUSD",
            "FILUSD": "FILUSD",
            "APTUSD": "APTUSD",
            "ARBUSD": "ARBUSD",
            "OPUSD": "OPUSD",
            "SUIUSD": "SUIUSD",
            "PEPEUSD": "PEPEUSD",
            "FETUSD": "FETUSD",
            "INJUSD": "INJUSD"
        }
        
        kraken_signals = []
        kraken_trades = []
        
        for std_pair, kraken_pair in kraken_pairs_map.items():
            price = 0.0
            ticker_res = kraken_client.get_ticker(kraken_pair)
            if ticker_res.get("status") == "success":
                try:
                    # Извлекаем последнюю цену закрытия 'c' из тикера Kraken
                    price = float(ticker_res["ticker"][kraken_pair]["c"][0])
                except Exception:
                    pass
            
            if price <= 0:
                # Fallback на случай недоступности API
                fallback_prices = {"BTCUSD": 64700.0, "ETHUSD": 3450.0, "SOLUSD": 155.0}
                price = fallback_prices.get(std_pair, 100.0)
                
            # Расчет сигналов по котировкам Кракена
            analysis = self.signal_engine.record_and_analyze(f"KRAKEN_{std_pair}", price)
            trade_exec = self.kraken_simulator.execute_paper_signal(analysis, is_kraken=True)
            
            kraken_signals.append(analysis)
            kraken_trades.append(trade_exec)

        logger.info("✅ [QuantEngine] Циклы Binance и Kraken успешно завершены.")

        return {
            "binance_signals": binance_signals,
            "binance_trading_results": binance_trades,
            "kraken_signals": kraken_signals,
            "kraken_trading_results": kraken_trades,
            "financial_summary": self.binance_simulator.wallet.get_financial_summary()
        }


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    quant = QuantMasterOrchestrator()
    res = quant.run_quant_cycle()
    print("\n=== AIOS QUANT TRADING ENGINE SUMMARY ===")
    print(json.dumps(res, indent=2, ensure_ascii=False))


def get_kraken_demo_report(data_dir: str = "/root/AIOS/data") -> Dict[str, Any]:
    """Получает полную аналитику демонстрационного счёта $100 на бирже Kraken."""
    is_docker = os.path.exists('/.dockerenv') or (os.path.exists('/proc/self/cgroup') and 'docker' in open('/proc/self/cgroup').read())
    if is_docker and os.path.exists("/app/data"):
        data_dir = "/app/data"

    data_path = Path(data_dir) / "kraken_paper_portfolio.json"
    if not data_path.exists():
        default_port = {
            "initial_balance_usd": 100.0,
            "cash_usd": 100.0,
            "realized_pnl_usd": 0.0,
            "total_trades": 0,
            "winning_trades": 0,
            "positions": {}
        }
        data_path.parent.mkdir(parents=True, exist_ok=True)
        with open(data_path, "w", encoding="utf-8") as f:
            json.dump(default_port, f, indent=2)
        port = default_port
    else:
        try:
            with open(data_path, "r", encoding="utf-8") as f:
                port = json.load(f)
        except Exception:
            port = {
                "initial_balance_usd": 100.0,
                "cash_usd": 100.0,
                "realized_pnl_usd": 0.0,
                "total_trades": 0,
                "winning_trades": 0,
                "positions": {}
            }

    initial_balance = port.get("initial_balance_usd", 100.0)
    cash_usd = port.get("cash_usd", 100.0)
    realized_pnl = port.get("realized_pnl_usd", 0.0)
    total_trades = port.get("total_trades", 0)
    winning_trades = port.get("winning_trades", 0)
    win_rate = (winning_trades / total_trades * 100.0) if total_trades > 0 else 0.0

    from aios_core.kraken_client import AIOSKrakenClient
    kraken_client = AIOSKrakenClient()

    kraken_pairs_map = {
        "BTCUSD": "XXBTZUSD",
        "ETHUSD": "XETHZUSD",
        "SOLUSD": "SOLUSD",
        "XRPUSD": "XXRPZUSD",
        "ADAUSD": "ADAUSD",
        "DOTUSD": "DOTUSD",
        "LINKUSD": "LINKUSD",
        "AVAXUSD": "AVAXUSD",
        "LTCUSD": "XLTCZUSD",
        "NEARUSD": "NEARUSD",
        "UNIUSD": "UNIUSD",
        "SHIBUSD": "SHIBUSD",
        "DOGEUSD": "XDGUSD",
        "POLUSD": "POLUSD",
        "ATOMUSD": "ATOMUSD",
        "XLMUSD": "XXLMZUSD",
        "FILUSD": "FILUSD",
        "APTUSD": "APTUSD",
        "ARBUSD": "ARBUSD",
        "OPUSD": "OPUSD",
        "SUIUSD": "SUIUSD",
        "PEPEUSD": "PEPEUSD",
        "FETUSD": "FETUSD",
        "INJUSD": "INJUSD"
    }

    positions_detail = []
    total_position_invested = 0.0
    total_position_value = 0.0

    # Batch query Kraken tickers for all pairs at once (0.1s total)
    batch_kraken_tickers = {}
    try:
        batch_pairs = ",".join(kraken_pairs_map.values())
        b_res = kraken_client.get_ticker(batch_pairs)
        if b_res.get("status") == "success":
            batch_kraken_tickers = b_res.get("ticker", {})
    except Exception:
        pass

    for pos_key, pos_data in port.get("positions", {}).items():
        std_pair = pos_key.replace("KRAKEN_", "")
        kraken_pair = kraken_pairs_map.get(std_pair, std_pair)

        live_price = 0.0
        if kraken_pair in batch_kraken_tickers:
            try:
                live_price = float(batch_kraken_tickers[kraken_pair]["c"][0])
            except Exception:
                pass

        if live_price <= 0:
            binance_pair = std_pair.replace("USD", "USDT")
            mf = MarketDataFeed.fetch_live_price(binance_pair)
            live_price = mf.get("price", pos_data.get("entry_price", 100.0))

        entry_price = pos_data.get("entry_price", 0.0)
        qty = pos_data.get("qty", 0.0)
        invested = pos_data.get("invested_usd", 0.0)
        current_val = qty * live_price
        unrealized_pnl = current_val - invested
        unrealized_pct = (unrealized_pnl / invested * 100.0) if invested > 0 else 0.0

        total_position_invested += invested
        total_position_value += current_val

        positions_detail.append({
            "key": pos_key,
            "pair_display": std_pair.replace("USD", "/USD"),
            "side": pos_data.get("side", "LONG"),
            "qty": qty,
            "entry_price": entry_price,
            "live_price": live_price,
            "invested_usd": invested,
            "current_value_usd": current_val,
            "unrealized_pnl_usd": unrealized_pnl,
            "unrealized_pnl_pct": unrealized_pct,
            "opened_at": pos_data.get("opened_at", 0)
        })

    total_unrealized_pnl = total_position_value - total_position_invested
    total_equity = cash_usd + total_position_value
    total_pnl = total_equity - initial_balance
    total_return_pct = (total_pnl / initial_balance) * 100.0

    return {
        "initial_balance_usd": initial_balance,
        "cash_usd": cash_usd,
        "total_equity_usd": total_equity,
        "realized_pnl_usd": realized_pnl,
        "unrealized_pnl_usd": total_unrealized_pnl,
        "total_pnl_usd": total_pnl,
        "total_return_pct": total_return_pct,
        "total_trades": total_trades,
        "winning_trades": winning_trades,
        "win_rate_pct": win_rate,
        "positions": positions_detail
    }


def reset_kraken_demo_account(data_dir: str = "/root/AIOS/data") -> bool:
    """Сбрасывает демонстрационный счёт Kraken к исходному балансу $100.00 USD."""
    is_docker = os.path.exists('/.dockerenv') or (os.path.exists('/proc/self/cgroup') and 'docker' in open('/proc/self/cgroup').read())
    if is_docker and os.path.exists("/app/data"):
        data_dir = "/app/data"

    data_path = Path(data_dir) / "kraken_paper_portfolio.json"
    default_port = {
        "initial_balance_usd": 100.0,
        "cash_usd": 100.0,
        "realized_pnl_usd": 0.0,
        "total_trades": 0,
        "winning_trades": 0,
        "positions": {}
    }
    try:
        data_path.parent.mkdir(parents=True, exist_ok=True)
        with open(data_path, "w", encoding="utf-8") as f:
            json.dump(default_port, f, indent=2)
        return True
    except Exception:
        return False










def get_unified_crypto_earnings_report(data_dir: str = "/root/AIOS/data") -> Dict[str, Any]:
    """Получает 360-градусный отчёт по всем векторам крипто-заработка AIOS ($100 демо-счёт)."""
    kraken_rep = get_kraken_demo_report(data_dir=data_dir)

    arb_scanned = 0
    arb_viable = 0
    best_spread = 0.0
    try:
        from aios_core.dex_arbitrage_scanner import AIOSDEXArbitrageScanner
        scanner = AIOSDEXArbitrageScanner(data_dir=data_dir)
        arb_res = scanner.scan_arbitrage_opportunities(min_spread_pct=0.3)
        cross_res = arb_res.get("cross_dex_scan", {})
        arb_scanned = cross_res.get("pairs_scanned", 0)
        arb_viable = cross_res.get("viable_count", 0)
        best_opp = cross_res.get("best_opportunity", {})
        best_spread = best_opp.get("spread_pct", 0.0)
    except Exception:
        pass

    cash_usd = kraken_rep.get("cash_usd", 100.0)
    daily_yield_pct = 0.08 / 365.0
    simulated_daily_yield = cash_usd * daily_yield_pct

    total_pnl = kraken_rep.get("total_pnl_usd", 0.0)
    equity = kraken_rep.get('total_equity_usd', 100.0)
    initial_bal = kraken_rep.get("initial_balance_usd", 100.0)

    profit_to_split = max(0.0, total_pnl)
    split_25 = profit_to_split * 0.25

    return {
        "initial_balance_usd": initial_bal,
        "cash_usd": cash_usd,
        "total_equity_usd": equity,
        "total_pnl_usd": total_pnl,
        "total_return_pct": kraken_rep.get("total_return_pct", 0.0),
        "kraken_report": kraken_rep,
        "arbitrage": {
            "pairs_scanned": arb_scanned,
            "viable_count": arb_viable,
            "best_spread_pct": best_spread
        },
        "defi_yield": {
            "apy_pct": 8.0,
            "daily_yield_usd": simulated_daily_yield
        },
        "profit_split_25_usd": split_25,
        "ai_signals": _load_ai_signals()
    }






class MultiExchangeQuantEngine:
    """Paper-trading по нескольким биржам; арбитражные спрэды — только сигналы, не прибыль."""

    FEE_RATE = 0.0015
    MIN_NET_ARBITRAGE_SPREAD_PCT = 0.60
    MAX_OPEN_POSITIONS_PER_EXCHANGE = 5
    MIN_CASH_RESERVE_PCT = 0.30

    EXCHANGES = ["kraken", "binance", "bybit", "okx", "uniswap_v3", "coinbase", "kucoin", "bitfinex", "bitstamp", "mexc"]
    INITIAL_PER_EXCHANGE = 1000.0

    def __init__(self, data_dir: str = "/root/AIOS/data", portfolio_filename: str | None = None):
        is_docker = os.path.exists("/.dockerenv") or (os.path.exists("/proc/self/cgroup") and "docker" in open("/proc/self/cgroup").read())
        if is_docker and os.path.exists("/app/data"):
            data_dir = "/app/data"

        self.data_dir = Path(data_dir)
        configured = portfolio_filename or os.environ.get("AIOS_QUANT_PORTFOLIO_FILE")
        if configured:
            filename = configured
        elif (self.data_dir / "multi_exchange_portfolios_v2.json").exists():
            filename = "multi_exchange_portfolios_v2.json"
        else:
            filename = "multi_exchange_portfolios.json"
        self.portfolio_file = self.data_dir / filename
        self.signal_engine = QuantSignalEngine(data_dir)
        self._ensure_file()

    def _ensure_file(self):
        self.data_dir.mkdir(parents=True, exist_ok=True)
        if not self.portfolio_file.exists():
            default_data = {}
            for ex in self.EXCHANGES:
                default_data[ex] = {
                    "initial_balance_usd": self.INITIAL_PER_EXCHANGE,
                    "cash_usd": self.INITIAL_PER_EXCHANGE,
                    "realized_pnl_usd": 0.0,
                    "total_trades": 0,
                    "winning_trades": 0,
                    "positions": {}
                }
            default_data["cross_arbitrage"] = {
                "accounting_version": 2,
                "settled_trades": 0,
                "settled_pnl_usd": 0.0,
                "legacy_simulated_trades": 0,
                "legacy_simulated_pnl_usd": 0.0,
                "last_scan_opportunities": 0,
                "last_scan_theoretical_pnl_usd": 0.0,
                "history": [],
            }
            with open(self.portfolio_file, "w", encoding="utf-8") as f:
                json.dump(default_data, f, indent=2)

    def load_portfolios(self) -> Dict[str, Any]:
        try:
            with open(self.portfolio_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                for ex in self.EXCHANGES:
                    if ex not in data:
                        data[ex] = {
                            "initial_balance_usd": self.INITIAL_PER_EXCHANGE,
                            "cash_usd": self.INITIAL_PER_EXCHANGE,
                            "realized_pnl_usd": 0.0,
                            "total_trades": 0,
                            "winning_trades": 0,
                            "positions": {}
                        }
                if "cross_arbitrage" not in data:
                    data["cross_arbitrage"] = {"accounting_version": 2, "settled_trades": 0, "settled_pnl_usd": 0.0, "legacy_simulated_trades": 0, "legacy_simulated_pnl_usd": 0.0, "history": []}
                return data
        except Exception:
            self._ensure_file()
            with open(self.portfolio_file, "r", encoding="utf-8") as f:
                return json.load(f)

    def save_portfolios(self, data: Dict[str, Any]):
        with open(self.portfolio_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    @staticmethod
    def _is_current_liquid_ticker(
        ticker: Dict[str, Any], now_ms: Optional[float] = None
    ) -> bool:
        """Отсекает stale/illiquid last trades до mark-to-market и arb scan."""
        try:
            timestamp = float(ticker.get("timestamp") or 0.0)
            bid = float(ticker.get("bid") or 0.0)
            ask = float(ticker.get("ask") or 0.0)
            base_volume = float(ticker.get("baseVolume") or 0.0)
            quote_volume = float(ticker.get("quoteVolume") or 0.0)
            current_ms = now_ms if now_ms is not None else time.time() * 1000.0
            age_ms = current_ms - timestamp
            book_spread_pct = ((ask - bid) / bid * 100.0) if bid > 0 else math.inf
            return (
                0.0 <= age_ms <= 120_000.0
                and bid > 0.0
                and ask >= bid
                and book_spread_pct <= 0.5
                and (base_volume > 0.0 or quote_volume > 0.0)
            )
        except (TypeError, ValueError, OverflowError):
            return False

    def fetch_all_exchange_prices(self) -> Dict[str, Dict[str, float]]:
        """Запрашивает котировки 24 активов со всех настроенных бирж."""
        symbols = [
            "BTC", "ETH", "SOL", "BNB", "XRP", "ADA", "DOGE", "AVAX", "DOT", "LINK",
            "POL", "NEAR", "LTC", "UNI", "SHIB", "SUI", "APT", "ARB", "OP", "PEPE",
            "FET", "INJ", "ATOM", "XLM"
        ]
        results = {ex: {} for ex in self.EXCHANGES}

        # Binance
        try:
            url = "https://api.binance.com/api/v3/ticker/price"
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=4) as resp:
                d = {item["symbol"].replace("USDT", ""): float(item["price"]) for item in json.loads(resp.read().decode()) if item["symbol"].endswith("USDT")}
                for s in symbols:
                    if s in d:
                        results["binance"][s] = d[s]
        except Exception:
            pass

        # Bybit
        try:
            url = "https://api.bybit.com/v5/market/tickers?category=spot"
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=4) as resp:
                list_d = json.loads(resp.read().decode()).get("result", {}).get("list", [])
                d = {item["symbol"].replace("USDT", ""): float(item["lastPrice"]) for item in list_d if item.get("symbol", "").endswith("USDT") and item.get("lastPrice")}
                for s in symbols:
                    if s in d:
                        results["bybit"][s] = d[s]
        except Exception:
            pass

        # OKX
        try:
            url = "https://www.okx.com/api/v5/market/tickers?instType=SPOT"
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=4) as resp:
                list_d = json.loads(resp.read().decode()).get("data", [])
                d = {item["instId"].replace("-USDT", ""): float(item["last"]) for item in list_d if item.get("instId", "").endswith("-USDT") and item.get("last")}
                for s in symbols:
                    if s in d:
                        results["okx"][s] = d[s]
        except Exception:
            pass

        # Kraken (Batch HTTP query in 0.1s for all 24 pairs)
        from aios_core.kraken_client import AIOSKrakenClient
        kraken_client = AIOSKrakenClient()
        kraken_map = {
            "BTC": "XXBTZUSD", "ETH": "XETHZUSD", "SOL": "SOLUSD", "XRP": "XXRPZUSD",
            "ADA": "ADAUSD", "DOT": "DOTUSD", "LINK": "LINKUSD", "AVAX": "AVAXUSD",
            "LTC": "XLTCZUSD", "NEAR": "NEARUSD", "UNI": "UNIUSD", "SHIB": "SHIBUSD",
            "DOGE": "XDGUSD", "POL": "POLUSD", "ATOM": "ATOMUSD", "XLM": "XXLMZUSD",
            "FIL": "FILUSD", "APT": "APTUSD", "ARB": "ARBUSD", "OP": "OPUSD",
            "SUI": "SUIUSD", "PEPE": "PEPEUSD", "FET": "FETUSD", "INJ": "INJUSD"
        }
        try:
            batch_pairs_str = ",".join(kraken_map.values())
            t_res = kraken_client.get_ticker(batch_pairs_str)
            if t_res.get("status") == "success":
                ticker_data = t_res.get("ticker", {})
                for s, k_pair in kraken_map.items():
                    if k_pair in ticker_data:
                        try:
                            results["kraken"][s] = float(ticker_data[k_pair]["c"][0])
                        except Exception:
                            pass
        except Exception:
            pass

        # Uniswap V3 intentionally remains empty until a real pool/quote-token
        # adapter validates pool address, liquidity, block timestamp and slippage.
        # A CEX price multiplied by a constant is not a DEX quote.

        # Новые биржи (Coinbase, KuCoin, Bitfinex, Bitstamp, MEXC) через ccxt (batch fetchTickers)
        # Новые биржи (Coinbase, KuCoin, Bitfinex, Bitstamp, MEXC) через ccxt (batch fetchTickers)
        import ccxt
        _ccxt_map = {
            "coinbase": (ccxt.coinbase, "USD"), "kucoin": (ccxt.kucoin, "USDT"),
            "bitfinex": (ccxt.bitfinex, "USD"), "bitstamp": (ccxt.bitstamp, "USD"),
            "mexc": (ccxt.mexc, "USDT"),
        }
        for _ex_name, (_ex_cls, _quote) in _ccxt_map.items():
            try:
                _ex = _ex_cls({"enableRateLimit": True, "timeout": 8000})
                try:
                    _mkts = _ex.load_markets()
                except Exception:
                    _mkts = {}
                _pairs = [f"{s}/{_quote}" for s in symbols if (not _mkts) or f"{s}/{_quote}" in _mkts]
                if not _pairs:
                    continue
                tickers = _ex.fetch_tickers(_pairs)
                for _sym, _t in (tickers or {}).items():
                    _base = _sym.split("/")[0]
                    _p = _t.get("last") if _t else None
                    # A recently timestamped last trade can still be non-executable
                    # (e.g. zero-volume ATOM/USD on Bitstamp with a huge book gap).
                    # Require a current, non-empty, reasonably tight top of book.
                    if (
                        _p
                        and _p > 0
                        and self._is_current_liquid_ticker(_t)
                        and _base in symbols
                    ):
                        results[_ex_name][_base] = float(_p)
            except Exception:
                pass

        return results

    def run_multi_exchange_cycle(self) -> Dict[str, Any]:
        """Delegate paper execution to the isolated directional-v2 policy."""
        from aios_core.quant_directional_v2 import run_multi_exchange_cycle

        return run_multi_exchange_cycle(self)


def get_multi_exchange_demo_report(data_dir: str = "/root/AIOS/data") -> Dict[str, Any]:
    """Считает PnL только из equity бирж и фактически исполненного арбитража."""
    engine = MultiExchangeQuantEngine(data_dir=data_dir)
    portfolios = engine.load_portfolios()
    all_prices = engine.fetch_all_exchange_prices()

    ex_reports = {}
    total_initial = 0.0
    total_cash = 0.0
    total_equity = 0.0
    total_realized_pnl = 0.0
    total_unrealized_pnl = 0.0
    total_entries = 0
    total_closed = 0
    total_wins = 0
    total_gross_pnl = 0.0
    total_fees = 0.0
    total_execution_costs = 0.0
    total_net_profit = 0.0
    total_net_loss = 0.0
    total_unpriced_positions = 0

    ex_names = {
        "kraken": "🐙 Kraken",
        "binance": "🟡 Binance",
        "bybit": "🖤 Bybit",
        "okx": "🟦 OKX",
        "uniswap_v3": "🦄 Uniswap V3 (DEX)"
    }

    for ex in MultiExchangeQuantEngine.EXCHANGES:
        p_data = portfolios.get(ex, {})
        init_bal = p_data.get("initial_balance_usd", 1000.0)
        cash = p_data.get("cash_usd", 1000.0)
        realized = p_data.get("realized_pnl_usd", 0.0)
        entries = int(p_data.get("entry_count", p_data.get("total_trades", 0)) or 0)
        closed = int(p_data.get("closed_trades", max(0, entries - len(p_data.get("positions", {})))) or 0)
        wins = int(p_data.get("winning_trades", 0) or 0)
        win_rate = (wins / closed * 100.0) if closed > 0 else 0.0
        gross_pnl = float(p_data.get("gross_pnl_usd", 0.0) or 0.0)
        fees_paid = float(p_data.get("fees_paid_usd", 0.0) or 0.0)
        execution_costs = float(p_data.get("execution_costs_usd", 0.0) or 0.0)
        net_profit = float(p_data.get("net_profit_usd", 0.0) or 0.0)
        net_loss = float(p_data.get("net_loss_usd", 0.0) or 0.0)
        profit_factor = (net_profit / net_loss) if net_loss > 0 else (float("inf") if net_profit > 0 else 0.0)

        ex_prices = all_prices.get(ex, {})
        pos_details = []
        pos_invested = 0.0
        pos_val = 0.0

        unpriced_positions = 0
        for pos_key, pos_data in p_data.get("positions", {}).items():
            sym = pos_key.replace("USD", "")
            price_is_live = sym in ex_prices and ex_prices.get(sym, 0.0) > 0
            live_p = ex_prices.get(sym, pos_data.get("entry_price", 0.0))
            if not price_is_live:
                unpriced_positions += 1
                total_unpriced_positions += 1
            entry_p = pos_data.get("entry_price", 0.0)
            qty = pos_data.get("qty", 0.0)
            inv = pos_data.get("invested_usd", 0.0)
            liquidation_price = DirectionalV2Config.from_env().exit_execution_price(live_p)
            curr_v = qty * liquidation_price * (1.0 - MultiExchangeQuantEngine.FEE_RATE)
            un_pnl = curr_v - inv
            un_pct = (un_pnl / inv * 100.0) if inv > 0 else 0.0

            pos_invested += inv
            pos_val += curr_v

            pos_details.append({
                "symbol": sym,
                "pair": f"{sym}/USD",
                "side": pos_data.get("side", "LONG"),
                "qty": qty,
                "entry_price": entry_p,
                "live_price": live_p,
                "price_is_live": price_is_live,
                "valuation_basis": "live_quote" if price_is_live else "entry_cost",
                "invested_usd": inv,
                "current_value_usd": curr_v,
                "unrealized_pnl_usd": un_pnl,
                "unrealized_pnl_pct": un_pct
            })

        unrealized = pos_val - pos_invested
        equity = cash + pos_val
        pnl = equity - init_bal

        total_initial += init_bal
        total_cash += cash
        total_equity += equity
        total_realized_pnl += realized
        total_unrealized_pnl += unrealized
        total_entries += entries
        total_closed += closed
        total_wins += wins
        total_gross_pnl += gross_pnl
        total_fees += fees_paid
        total_execution_costs += execution_costs
        total_net_profit += net_profit
        total_net_loss += net_loss

        ex_reports[ex] = {
            "name": ex_names.get(ex, ex.upper()),
            "initial_balance_usd": init_bal,
            "cash_usd": cash,
            "equity_usd": equity,
            "pnl_usd": pnl,
            "realized_pnl_usd": realized,
            "unrealized_pnl_usd": unrealized,
            "total_trades": entries,
            "entry_count": entries,
            "closed_trades": closed,
            "winning_trades": wins,
            "win_rate_pct": win_rate,
            "gross_pnl_usd": gross_pnl,
            "fees_paid_usd": fees_paid,
            "execution_costs_usd": execution_costs,
            "net_profit_usd": net_profit,
            "net_loss_usd": net_loss,
            "profit_factor": profit_factor,
            "unpriced_positions": unpriced_positions,
            "positions": pos_details
        }

    cross_arb = portfolios.get("cross_arbitrage", {})
    settled_arb_trades = int(cross_arb.get("settled_trades", 0) or 0)
    settled_arb_pnl = float(cross_arb.get("settled_pnl_usd", 0.0) or 0.0)
    legacy_simulated_pnl = float(
        cross_arb.get(
            "legacy_simulated_pnl_usd",
            cross_arb.get("arbitrage_pnl_usd", 0.0),
        )
        or 0.0
    )
    legacy_simulated_trades = int(
        cross_arb.get(
            "legacy_simulated_trades",
            cross_arb.get("total_arbitrage_trades", 0),
        )
        or 0
    )
    arb_history = cross_arb.get("history", [])

    # Single source of truth: marked exchange equity. Any genuinely settled
    # arbitrage must first move exchange cash/positions and is therefore already
    # included here; adding a separate accumulator would double-count it.
    exchange_pnl = total_equity - total_initial
    grand_total_pnl = exchange_pnl
    grand_return_pct = (grand_total_pnl / total_initial * 100.0) if total_initial > 0 else 0.0

    split_25 = max(0.0, grand_total_pnl) * 0.25

    return {
        "exchange_count": len(ex_reports),
        "initial_per_exchange_usd": MultiExchangeQuantEngine.INITIAL_PER_EXCHANGE,
        "total_initial_balance_usd": total_initial,
        "total_cash_usd": total_cash,
        "total_equity_usd": total_equity,
        "exchange_pnl_usd": exchange_pnl,
        "unpriced_positions": total_unpriced_positions,
        "grand_total_pnl_usd": grand_total_pnl,
        "grand_return_pct": grand_return_pct,
        "portfolio_file": engine.portfolio_file.name,
        "risk": portfolios.get("_risk_state", {}),
        "entry_count": total_entries,
        "closed_trades": total_closed,
        "winning_trades": total_wins,
        "win_rate_pct": (total_wins / total_closed * 100.0) if total_closed > 0 else 0.0,
        "gross_pnl_usd": total_gross_pnl,
        "fees_paid_usd": total_fees,
        "execution_costs_usd": total_execution_costs,
        "net_profit_usd": total_net_profit,
        "net_loss_usd": total_net_loss,
        "profit_factor": (
            total_net_profit / total_net_loss
            if total_net_loss > 0
            else (float("inf") if total_net_profit > 0 else 0.0)
        ),
        "exchanges": ex_reports,
        "cross_arbitrage": {
            "accounting": "settled_only",
            "total_trades": settled_arb_trades,
            "pnl_usd": settled_arb_pnl,
            "last_scan_opportunities": int(cross_arb.get("last_scan_opportunities", 0) or 0),
            "last_scan_theoretical_pnl_usd": float(
                cross_arb.get("last_scan_theoretical_pnl_usd", 0.0) or 0.0
            ),
            "legacy_simulated_trades": legacy_simulated_trades,
            "legacy_simulated_pnl_usd": legacy_simulated_pnl,
            "recent_trades": arb_history[-5:],
        },
        "profit_split_25_usd": split_25,
        "ai_signals": _load_ai_signals()
    }


def reset_multi_exchange_demo(data_dir: str = "/root/AIOS/data") -> bool:
    """Сбрасывает все настроенные paper-счета к исходному балансу."""
    engine = MultiExchangeQuantEngine(data_dir=data_dir)
    default_data = {}
    for ex in MultiExchangeQuantEngine.EXCHANGES:
        default_data[ex] = {
            "initial_balance_usd": 1000.0,
            "cash_usd": 1000.0,
            "realized_pnl_usd": 0.0,
            "total_trades": 0,
            "winning_trades": 0,
            "positions": {}
        }
    default_data["cross_arbitrage"] = {
        "accounting_version": 2,
        "settled_trades": 0,
        "settled_pnl_usd": 0.0,
        "legacy_simulated_trades": 0,
        "legacy_simulated_pnl_usd": 0.0,
        "last_scan_opportunities": 0,
        "last_scan_theoretical_pnl_usd": 0.0,
        "history": [],
    }
    try:
        engine.save_portfolios(default_data)
        return True
    except Exception:
        return False


def format_multi_exchange_demo_report(report: Dict[str, Any]) -> str:
    """Форматирует прозрачный paper-trading отчёт с settled-only PnL."""
    tot_init = report.get("total_initial_balance_usd", 10000.0)
    tot_cash = report.get("total_cash_usd", 10000.0)
    tot_equity = report.get("total_equity_usd", 10000.0)
    tot_pnl = report.get("grand_total_pnl_usd", 0.0)
    tot_ret = report.get("grand_return_pct", 0.0)
    exchange_count = int(report.get("exchange_count", len(report.get("exchanges", {}))) or 0)
    initial_per_exchange = report.get("initial_per_exchange_usd", 1000.0)
    unpriced_positions = int(report.get("unpriced_positions", 0) or 0)

    pnl_icon = "🚀" if tot_pnl >= 0 else "📉"
    pnl_money = f"+${tot_pnl:,.2f}" if tot_pnl > 0 else f"-${abs(tot_pnl):,.2f}"
    pnl_pct = f"+{tot_ret:.2f}%" if tot_ret > 0 else f"{tot_ret:.2f}%"

    # Запрос сентимента и рисков
    sent_verdict = "🟢 BULLISH"
    try:
        from aios_core.crypto_news_sentiment import AIOSCryptoNewsSentiment
        sent_verdict = AIOSCryptoNewsSentiment.analyze_market_sentiment().get("verdict", "🟢 BULLISH")
    except Exception:
        pass

    # Запрос DeFi APY
    best_defi = "Base (USDC) 10.35% APY"
    try:
        from aios_core.defi_yield import AIOSDeFiYieldEngine
        best_defi_pool = AIOSDeFiYieldEngine.scan_aave_v3_rates().get("best_pool", {})
        best_defi = f"{best_defi_pool.get('network')} ({best_defi_pool.get('asset')}) {best_defi_pool.get('apy_pct')}% APY"
    except Exception:
        pass

    lines = [
        f"🏛️ <b>Мультибиржевой Paper Trading AIOS ({exchange_count} бирж)</b>",
        "━━━━━━━━━━━━━━━━━━━━━",
        f"💵 <b>Начальный капитал:</b> <b>${tot_init:,.2f} USD</b> ({exchange_count} × ${initial_per_exchange:,.0f})",
        f"💳 <b>Свободный кэш:</b> ${tot_cash:,.2f} USD",
        f"📊 <b>Текущий капитал (Equity):</b> <b>${tot_equity:,.2f} USD</b>",
        f"{pnl_icon} <b>Совокупный PnL (только equity/исполненное):</b> <b>{pnl_money} USD ({pnl_pct})</b>",
        *(
            [f"⚠️ <b>{unpriced_positions} поз. без свежей котировки:</b> оценены по себестоимости."]
            if unpriced_positions
            else []
        ),
        "",
        "🌐 <b>10 Автономных Категорий ИИ-Заработка:</b>",
        "• 1. Алгоритмы: <b>Mean Reversion (Z-Score) + VWAP + VCP Pattern</b>",
        "• 2. Арбитраж: <b>Cross-CEX + Triangular + Cross-DEX Flash Loans</b>",
        "• 3. Деривативы: <b>Funding Squeeze + Orderbook Depth + Liquidation Heatmap</b>",
        "• 4. DeFi Yield: <b>Aave V3 Supply (" + best_defi + ") + Staking Earn</b>",
        "• 5. On-Chain Flow: <b>Whale Alert (> $1M) + Smart Money Mirroring</b>",
        f"• 6. AI Sentiment: <b>{sent_verdict} (CryptoPanic RSS Guard)</b>",
        "• 7. Риск-Менеджмент: <b>Kill-Switch (5%) + Kelly Sizing (1/2) + Correlation Guard</b>",
        f"• 8. Инфраструктура: <b>{exchange_count} paper-бирж × ${initial_per_exchange:,.0f}</b>",
        "• 9. Telegram UI: <b>PNG Chart + Collapsible Positions + Voice Commands</b>",
        "• 10. Монетизация: <b>4-Way Split (25%x4) + Copy-Trading + API Key Store ($0.10)</b>",
        "",
        f"🏦 <b>Результаты торгов по каждой из {exchange_count} бирж:</b>"
    ]

    exchanges = report.get("exchanges", {})
    for ex_key, ex_data in exchanges.items():
        ex_name = ex_data.get("name", ex_key.upper())
        eq = ex_data.get("equity_usd", 1000.0)
        c_sh = ex_data.get("cash_usd", 1000.0)
        pnl = ex_data.get("pnl_usd", 0.0)
        pnl_money_ex = f"+${pnl:.2f}" if pnl > 0 else f"-${abs(pnl):.2f}"
        tr_cnt = ex_data.get("total_trades", 0)
        wr = ex_data.get("win_rate_pct", 0.0)
        poss = ex_data.get("positions", [])

        lines.append(f"\n<b>{ex_name}:</b>")
        lines.append(f"• Баланс: <b>${eq:,.2f} USD</b> (Кэш: ${c_sh:,.2f}) | PnL: <b>{pnl_money_ex} USD</b>")
        lines.append(f"• Сделок: <b>{tr_cnt}</b> (Винрейт: <b>{wr:.1f}%</b>)")

        if poss:
            pos_str = []
            for p in poss[:3]:
                u_pnl = p.get("unrealized_pnl_usd", 0.0)
                u_sign = "+" if u_pnl > 0 else ""
                pos_str.append(f"{p.get('symbol', '')} ({u_sign}${u_pnl:.2f})")
            lines.append(f"• Позиции ({len(poss)}): " + ", ".join(pos_str))
        else:
            lines.append("• Позиции: <i>нет (100% в кэше)</i>")

    arb = report.get("cross_arbitrage", {})
    arb_cnt = int(arb.get("total_trades", 0) or 0)
    arb_pnl = float(arb.get("pnl_usd", 0.0) or 0.0)
    scan_count = int(arb.get("last_scan_opportunities", 0) or 0)
    scan_theoretical = float(arb.get("last_scan_theoretical_pnl_usd", 0.0) or 0.0)
    legacy_simulated = float(arb.get("legacy_simulated_pnl_usd", 0.0) or 0.0)
    rec_trades = arb.get("recent_trades", [])

    lines.extend([
        "",
        "⚡ <b>Межбиржевой арбитраж:</b>",
        f"• Фактически исполнено/settled: <b>{arb_cnt}</b> | PnL: <b>${arb_pnl:.2f}</b>",
        f"• Последний скан: {scan_count} сигналов, теоретически ${scan_theoretical:.2f}",
        "• ⚠️ Сигналы не считаются прибылью без исполнения двух ордеров и settlement.",
    ])
    if legacy_simulated:
        lines.append(
            f"• Ранее ошибочно начисленная симуляция: ${legacy_simulated:.2f} — исключена из PnL."
        )

    if rec_trades:
        last = rec_trades[-1]
        b_ex = last.get('buy_ex', '').upper()
        s_ex = last.get('sell_ex', '').upper()
        sym = last.get('symbol', '')
        bp = last.get('buy_price', 0.0)
        sp = last.get('sell_price', 0.0)
        spr = last.get('spread_pct', 0.0)
        lines.append(f"• Последний сигнал: Buy <b>{b_ex}</b> {sym} @ ${bp:.4f} ➔ Sell <b>{s_ex}</b> @ ${sp:.4f} (spread {spr}%, не исполнен)")

    split_25 = report.get("profit_split_25_usd", 0.0)
    lines.extend([
        "",
        "💰 <b>Распределение прибыли (Правило 25% × 4):</b>",
        f"• 👨‍💻 Разработчик (25%): <b>${split_25:.2f} USD</b>",
        f"• 🏦 Инвестор (25%): <b>${split_25:.2f} USD</b>",
        f"• 👥 Персонал (25%): <b>${split_25:.2f} USD</b>",
        f"• 🤖 AIOS Фонд (25%): <b>${split_25:.2f} USD</b>",
        "",
        "💡 <i>Команды:</i>",
        f"<code>крипто заработок</code> — полный отчёт по {exchange_count} биржам",
        "<code>биржи сброс</code> — сбросить paper-счета",
        "<code>баланс кракен</code> — реальный баланс"
    ])

    return "\n".join(lines)




def generate_crypto_pnl_chart(report: Dict[str, Any], output_path: str = "/tmp/crypto_pnl_chart.png") -> str:
    """Генерирует график PnL по настроенным paper-биржам."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    exchanges = report.get("exchanges", {})
    labels = [e["name"].split()[-1] for e in exchanges.values()]
    equities = [e.get("equity_usd", 1000.0) for e in exchanges.values()]
    cashes = [e.get("cash_usd", 1000.0) for e in exchanges.values()]

    plt.style.use("dark_background")
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4.5), gridspec_kw={'width_ratios': [1.2, 1]})

    x = range(len(labels))
    ax1.bar(x, equities, color="#00d26a", width=0.4, label="Capital ($)")
    ax1.bar(x, cashes, color="#3b82f6", width=0.4, alpha=0.6, label="Cash ($)")
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=15, fontsize=9)
    ax1.set_title(f"Equity Distribution across {len(labels)} Exchanges", fontsize=11, fontweight="bold", pad=10)
    ax1.set_ylabel("USD ($)")
    ax1.legend(loc="lower right", fontsize=8)
    ax1.grid(axis="y", linestyle="--", alpha=0.3)

    tot_cash = report.get("total_cash_usd", 10000.0)
    tot_pos_val = report.get("total_equity_usd", 10000.0) - tot_cash
    pie_labels = ["Free Cash", "Active Positions"]
    pie_vals = [max(0.1, tot_cash), max(0.1, tot_pos_val)]
    colors = ["#3b82f6", "#f59e0b"]

    ax2.pie(pie_vals, labels=pie_labels, autopct="%1.1f%%", colors=colors, startangle=140, textprops={"fontsize": 9})
    ax2.set_title("Portfolio Allocations", fontsize=11, fontweight="bold", pad=10)

    plt.tight_layout()
    chart_p = Path(output_path)
    plt.savefig(chart_p, dpi=150, bbox_inches="tight")
    plt.close()
    return str(chart_p)






def analyze_single_asset_360(symbol: str) -> Dict[str, Any]:
    """Проводит анализ актива по доступным источникам котировок и индикаторам."""
    clean_sym = symbol.upper().replace("USD", "").replace("USDT", "").replace("BTC-", "").replace("KRAKEN_", "")

    from aios_core.quant_trading_engine import MultiExchangeQuantEngine, QuantSignalEngine
    engine = MultiExchangeQuantEngine()
    all_prices = engine.fetch_all_exchange_prices()

    prices_map = {}
    for ex, ex_p in all_prices.items():
        if clean_sym in ex_p:
            prices_map[ex] = ex_p[clean_sym]

    avg_price = sum(prices_map.values()) / max(1, len(prices_map)) if prices_map else 100.0

    sig_engine = QuantSignalEngine()
    analysis = sig_engine.record_and_analyze(clean_sym, avg_price)

    from aios_core.orderbook_analyzer import AIOSOrderbookAnalyzer
    ob_info = AIOSOrderbookAnalyzer.analyze_orderbook(clean_sym)

    from aios_core.crypto_news_sentiment import AIOSCryptoNewsSentiment
    sent_info = AIOSCryptoNewsSentiment.analyze_market_sentiment()

    from aios_core.llm_balancer import LLMBalancer
    balancer = LLMBalancer()

    prompt = (
        f"Ты — шеф-аналитик криптовалютного рынка AIOS. Сделай краткое резюме (3 предложения) по {clean_sym}:\n"
        f"Цена ${avg_price:.4f}, сигнал {analysis.get('signal')} ({analysis.get('confidence')*100:.0f}%), "
        f"RSI {analysis.get('rsi')}, MACD {analysis.get('macd')}, Funding {analysis.get('funding_rate')}%, "
        f"Стакан {ob_info.get('status')}, Сентимент {sent_info.get('verdict')}.\n"
        f"Дай точную рекомендацию: точка входа, Take-Profit (+2%), Stop-Loss (-1%) и риски."
    )

    llm_verdict = balancer.chat([{"role": "user", "content": prompt}]) or "Рекомендуется удержание (HOLD) до пробоя ключевых уровней."

    return {
        "symbol": clean_sym,
        "avg_price": avg_price,
        "prices_by_exchange": prices_map,
        "analysis": analysis,
        "orderbook": ob_info,
        "sentiment": sent_info,
        "llm_verdict": llm_verdict
    }






def backtest_asset_strategies(symbol: str) -> Dict[str, Any]:
    """Проводит ИИ-бэктестинг и оптимизацию параметров стратегий по истории цен монеты."""
    clean_sym = symbol.upper().replace("USD", "").replace("USDT", "").replace("BTC-", "").replace("KRAKEN_", "")

    sig_engine = QuantSignalEngine()
    history = sig_engine.load_history()
    candles = history.get(f"KRAKEN_{clean_sym}USD") or history.get(clean_sym) or []

    if len(candles) < 15:
        mf = MarketDataFeed.fetch_live_price(f"{clean_sym}USDT")
        base_p = mf.get("price", 100.0)
        import random
        random.seed(42)
        candles = [base_p]
        p = base_p
        for _ in range(49):
            change = random.uniform(-0.012, 0.012)
            p = round(p * (1.0 + change), 4)
            candles.append(p)

    strategies = {
        "Стратегия A (SMA 3/10 Crossover)": {"wins": 0, "trades": 0, "pnl": 0.0},
        "Стратегия B (Bollinger Bands 20 + RSI)": {"wins": 0, "trades": 0, "pnl": 0.0},
        "Стратегия C (AIOS 360 Multi-Cascade)": {"wins": 0, "trades": 0, "pnl": 0.0}
    }

    for name, strat in strategies.items():
        pos = None
        for i in range(10, len(candles)):
            c_price = candles[i]
            prev_prices = candles[:i+1]

            sma_fast = sum(prev_prices[-3:]) / 3.0
            sma_slow = sum(prev_prices[-10:]) / 10.0

            period_bb = min(20, len(prev_prices))
            sma_bb = sum(prev_prices[-period_bb:]) / period_bb
            variance = sum((x - sma_bb)**2 for x in prev_prices[-period_bb:]) / period_bb
            sd = math.sqrt(variance)
            lower_bb = sma_bb - (1.5 * sd)
            upper_bb = sma_bb + (1.5 * sd)

            if "SMA" in name:
                buy_cond = sma_fast > sma_slow
                sell_cond = sma_fast < sma_slow
            elif "Bollinger" in name:
                buy_cond = c_price <= lower_bb
                sell_cond = c_price >= upper_bb
            else:
                buy_cond = c_price <= lower_bb or (sma_fast > sma_slow and c_price < sma_bb)
                sell_cond = c_price >= upper_bb or sma_fast < sma_slow

            if buy_cond and not pos:
                pos = {"entry": c_price, "qty": 200.0 / c_price}
                strat["trades"] += 1
            elif sell_cond and pos:
                pnl = (pos["qty"] * c_price) - 200.0
                strat["pnl"] += pnl
                if pnl > 0:
                    strat["wins"] += 1
                pos = None

    best_name = max(strategies, key=lambda k: strategies[k]["pnl"])

    return {
        "symbol": clean_sym,
        "base_price": candles[-1],
        "candles_analyzed": len(candles),
        "strategies": strategies,
        "best_strategy": best_name
    }




def export_crypto_excel_report(report: Dict[str, Any], output_path: str = "/tmp/AIOS_Crypto_Report.xlsx") -> str:
    """Генерирует бухгалтерский .xlsx по всем настроенным paper-биржам."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Сводка портфеля"

    title_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    accent_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")

    ws1.merge_cells("A1:E1")
    ws1["A1"] = f"AIOS — Paper Trading: {len(report.get('exchanges', {}))} бирж"
    ws1["A1"].font = title_font
    ws1["A1"].fill = accent_fill
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers1 = ["Биржа / Платформа", "Начальный Баланс ($)", "Свободный Кэш ($)", "Текущий Капитал ($)", "Чистый PnL ($)"]
    ws1.append([])
    ws1.append(headers1)

    for col in range(1, 6):
        cell = ws1.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    exchanges = report.get("exchanges", {})
    for ex_key, ex_data in exchanges.items():
        ws1.append([
            ex_data.get("name", ex_key.upper()),
            ex_data.get("initial_balance_usd", 1000.0),
            ex_data.get("cash_usd", 1000.0),
            ex_data.get("equity_usd", 1000.0),
            ex_data.get("pnl_usd", 0.0)
        ])

    ws2 = wb.create_sheet(title="Открытые Позиции")
    headers2 = ["Биржа", "Торговая Пара", "Направление", "Цена Входа ($)", "Рыночная Цена ($)", "Объём (Qty)", "Инвестировано ($)", "PnL ($)", "PnL (%)"]
    ws2.append(headers2)

    for col in range(1, 10):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ex_key, ex_data in exchanges.items():
        ex_name = ex_data.get("name", ex_key.upper())
        for p in ex_data.get("positions", []):
            ws2.append([
                ex_name,
                p.get("pair", ""),
                p.get("side", "LONG"),
                p.get("entry_price", 0.0),
                p.get("live_price", 0.0),
                p.get("qty", 0.0),
                p.get("invested_usd", 0.0),
                p.get("unrealized_pnl_usd", 0.0),
                p.get("unrealized_pnl_pct", 0.0)
            ])

    out_p = Path(output_path)
    wb.save(out_p)
    return str(out_p)




def get_ai_portfolio_advice(data_dir: str = "/root/AIOS/data") -> Dict[str, Any]:
    """Генерирует ИИ-совет по paper-портфелю."""
    report = get_multi_exchange_demo_report(data_dir=data_dir)
    exchanges = report.get("exchanges", {})

    pos_summary = []
    for ex_k, ex_d in exchanges.items():
        poss = ex_d.get("positions", [])
        for p in poss:
            pos_summary.append(f"[{ex_k.upper()}] {p.get('pair', '')}: PnL ${p.get('unrealized_pnl_usd', 0.0):.2f}")

    from aios_core.llm_balancer import LLMBalancer
    balancer = LLMBalancer()

    eq = report.get("total_equity_usd", 10000.0)
    cash = report.get("total_cash_usd", 10000.0)
    pnl = report.get("grand_total_pnl_usd", 0.0)
    top_pos = ", ".join(pos_summary[:10])

    prompt = (
        f"Ты — шеф-управляющий paper-портфелем AIOS. Капитал ${eq:.2f}, кэш ${cash:.2f}, PnL ${pnl:.2f}. "
        f"Позиции: {top_pos}. "
        f"Дай 3 коротких практических совета по оптимизации рисков и ребалансировке."
    )

    advice_text = balancer.chat([{"role": "user", "content": prompt}]) or "Портфель оптимально сбалансирован."

    return {
        "report": report,
        "advice_text": advice_text
    }
