"""
AIOS Autonomous Financial Accountant & OOXML Excel Reporter
Модуль генерации профессиональных финансовых отчетов в формате Excel (.xlsx) на основе Gross-книги AIOS.
"""
from __future__ import annotations

import os
import json
import time
import logging
from pathlib import Path
from typing import Dict, Any, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from aios_core.crypto_wallet import AIOSWalletManager

logger = logging.getLogger("AIOS.AccountingReporter")


class AIOSAccountingReporter:
    """Автономный ИИ-Бухгалтер системы, собирающий финансовые балансы в Excel."""

    def __init__(self, data_dir: str = "/root/AIOS/data"):
        # Умное разрешение путей (Docker/Host)
        is_docker = os.path.exists('/.dockerenv') or (os.path.exists('/proc/self/cgroup') and 'docker' in open('/proc/self/cgroup').read())
        if is_docker and os.path.exists("/app/data"):
            data_dir = "/app/data"
            
        self.wallet_mgr = AIOSWalletManager(data_dir)
        self.data_dir = Path(data_dir)

    def generate_excel_report(self, output_path: str = "/root/AIOS/data/aios_financial_report.xlsx") -> str:
        """Генерирует красивый, профессионально оформленный финансовый отчет в формате .xlsx."""
        ledger = self.wallet_mgr.load_ledger()
        summary = self.wallet_mgr.get_financial_summary()
        
        # Получаем данные о невыплаченных On-Chain долгах из ledger
        unpaid = ledger.get("unpaid_shares_usd", {
            "developer": 0.0,
            "investor": 0.0,
            "personnel": 0.0
        })
        paid = ledger.get("paid_shares_usd", {
            "developer": 0.0,
            "investor": 0.0,
            "personnel": 0.0
        })

        wb = openpyxl.Workbook()
        
        # Стили оформления
        font_title = Font(name="Arial", size=14, bold=True, color="1F4E78")
        font_subtitle = Font(name="Arial", size=10, italic=True, color="595959")
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Arial", size=10, bold=False)
        font_bold_body = Font(name="Arial", size=10, bold=True)
        
        fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        fill_accent = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        thin_border_side = Side(style='thin', color='BFBFBF')
        border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # -------------------------------------------------------------
        # Лист 1: Сводка Баланса (Balance Summary)
        # -------------------------------------------------------------
        ws_balance = wb.active
        ws_balance.title = "Балансовый Отчет"
        ws_balance.views.sheetView[0].showGridLines = True
        
        ws_balance["A1"] = "ФИНАНСОВЫЙ БАЛАНС СИСТЕМЫ AIOS"
        ws_balance["A1"].font = font_title
        ws_balance["A2"] = f"Сформировано автоматически на дату: {time.strftime('%Y-%m-%d %H:%M:%S')}"
        ws_balance["A2"].font = font_subtitle
        
        # Таблица кошельков
        headers = ["Статья / Бюджетный кошелек", "Накоплено за все время (USD)", "Выплачено On-Chain (USD)", "Остаток к выплате (USD)"]
        for col_idx, text in enumerate(headers, 1):
            cell = ws_balance.cell(row=4, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_cell
            
        wallets_data = [
            ("1. Доля Разработчика (25%)", summary["wallet_balances_usd"]["1_developer_25pct"], paid.get("developer", 0.0), unpaid.get("developer", 0.0)),
            ("2. Доля Инвестора (25%)", summary["wallet_balances_usd"]["2_investor_25pct"], paid.get("investor", 0.0), unpaid.get("investor", 0.0)),
            ("3. Доля Персонала/Команды (25%)", summary["wallet_balances_usd"]["3_personnel_25pct"], paid.get("personnel", 0.0), unpaid.get("personnel", 0.0)),
            ("4. Автономный бюджет Системы (25%)", summary["wallet_balances_usd"]["4_system_autonomous_25pct"], 0.0, summary["wallet_balances_usd"]["4_system_autonomous_25pct"]),
        ]
        
        for row_idx, row_data in enumerate(wallets_data, 5):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws_balance.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_body
                cell.border = border_cell
                if col_idx > 1:
                    cell.number_format = "$#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                    
        # Итоговая строка активов
        total_earned = float(summary.get("total_earned_all_time_usd", 0.0))
        total_paid = sum(paid.values())
        total_remaining = total_earned - total_paid
        
        ws_balance.cell(row=9, column=1, value="ИТОГО АКТИВЫ").font = font_bold_body
        ws_balance.cell(row=9, column=1).fill = fill_total
        ws_balance.cell(row=9, column=1).border = border_cell
        
        for col_idx, val in enumerate([total_earned, total_paid, total_remaining], 2):
            cell = ws_balance.cell(row=9, column=col_idx, value=val)
            cell.font = font_bold_body
            cell.fill = fill_total
            cell.border = border_cell
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")

        # -------------------------------------------------------------
        # Лист 2: Журнал Операций (Transactions Log)
        # -------------------------------------------------------------
        ws_tx = wb.create_sheet(title="Журнал Операций")
        ws_tx.views.sheetView[0].showGridLines = True
        
        ws_tx["A1"] = "ПОЛНЫЙ РЕЕСТР ФИНАНСОВЫХ ОПЕРАЦИЙ"
        ws_tx["A1"].font = font_title
        ws_tx["A2"] = "Хронологический список зачислений, списаний расходов и авто-распределений"
        ws_tx["A2"].font = font_subtitle
        
        tx_headers = ["Дата / Время", "Тип Операции", "Сумма (USD)", "Источник / Назначение", "ID Задачи / TxHash"]
        for col_idx, text in enumerate(tx_headers, 1):
            cell = ws_tx.cell(row=4, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_cell
            
        transactions = ledger.get("transactions", [])
        for row_idx, tx in enumerate(reversed(transactions), 5): # Новые транзакции сверху
            dt_str = tx.get("datetime", "")
            tx_type = tx.get("type", "")
            amount = tx.get("total_amount_usd") or tx.get("amount_usd") or tx.get("amount") or 0.0
            source = tx.get("source") or tx.get("purpose") or tx.get("reason") or ""
            task_id = tx.get("task_id") or tx.get("tx_hash") or ""
            
            row_data = [dt_str, tx_type, amount, source, task_id]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws_tx.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_body
                cell.border = border_cell
                if col_idx == 3:
                    cell.number_format = "$#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                    
        # Автоматическая регулировка ширины колонок для всех листов
        for ws in [ws_balance, ws_tx]:
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.row in [1, 2]: # Пропускаем заголовки при расчете ширины
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        # Сохранение файла
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"📊 [Accounting] Финансовый отчет Excel успешно сгенерирован: {output_path}")
        return output_path
